"""Transcribe Annex I of Implementing Regulation (EU) 2025/2621 into two dbt seeds.

The CBAM default values are regulatory reference data: they are versioned by
amendment, not by scrape, so they belong in `dbt/seeds/` rather than behind a dlt
resource. This script exists so that the *next* amendment is a re-run rather than
a re-transcription — the Commission publishes an XLSX of the annex "for
information purposes only, while the legally binding values are set out in"
the regulation itself.

    uv run python -m scripts.build_cbam_seeds            # download, then write
    uv run python -m scripts.build_cbam_seeds --xlsx X   # from a local copy

Two seeds, because normalising the goods out is worth 1.6 MB. 12,540 value rows
share 283 distinct (product group, CN code, description) triples, and the
descriptions are long — one of them runs to 250 characters. Repeated per row they
are 1.6 MB of CSV and the same again in the warehouse table; as a dimension they
are 38 kB.

**The seeds now hold the annex as corrected by Implementing Regulation (EU)
2026/1740**, adopted 20 July 2026, in force 3 August and applying retroactively
from 1 January 2026. It replaces Annexes I and IV of 2025/2621 in full, and it is
the reason the amendment path above is worth having: re-running this script was
most of the migration. What it changed, because none of it is guessable from the
values alone —

- **The three mark-up columns are gone.** The February version published each
  good's value *including* the phase-in mark-up for 2026, 2027 and 2028 beside
  the plain total; this one publishes direct, indirect, total and route, and
  nothing else. The mark-up is unchanged in law — 10 / 20 / 30% for cement, iron
  and steel, aluminium and hydrogen, a flat 1% for fertilisers — so it moved from
  something this project *derived* to something it must *assert*. That is the
  `cbam_markup_schedule` seed.
- **10-digit TARIC codes appear**, and they resolve a real ambiguity: white and
  grey clinker were both `2523 10 00` and differ by more than 2x. They are now
  `2523 10 00 10` and `2523 10 00 90`.
- **Twelve sheet names changed** ("Russia" -> "Russian Federation", "Vietnam" ->
  "Viet Nam", ...) and two countries are new (Liberia, New Caledonia).
- **The values themselves barely moved**: 66 of 10,503 comparable rows, 38 of
  them down by 2% or less, 28 now blank.

What this script deliberately does *not* do is clean the numbers. The annex is
the legal instrument, so the seed is a faithful transcription of it, defects
included, and `fct_cbam_exposure` is where the published values meet the rules
that are supposed to govern them. The correction fixed two of the three defects
the February version had, which is worth recording because the handling that
absorbed them is now load-bearing for nothing:

- **Albania / 2523 21 00 (white Portland cement)** used to carry `–` for direct,
  indirect and total with 1,230 / 0,140 / 1,370 sitting in the *mark-up* columns
  instead — the values shifted three columns right. It is a clean `–` now.
- **Five cement rows** (Angola and Argentina) used to compound the mark-up —
  x1.1, x1.21, x1.331 — where the other 10,926 added it. With no published
  mark-up column there is nothing left to compound.
- **Fertilisers carry a 1% mark-up in all three years**, not 1/2/3% and not
  10/20/30%. Still true, still the reason the mark-up is a property of the
  product group rather than a constant, and now asserted rather than measured.

`dbt/seeds/_seeds.yml` and `dbt/models/marts/_marts.yml` test what remains, so
the next amendment shows up as a failing test rather than as a silent change in a
published euro figure.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
import unicodedata
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path
from urllib.request import urlopen

from modern_data_stack.paths import project_root

# "for information purposes only" per the Commission's own page; the binding text
# is the OJ. Kept here rather than in the docs because it is what the script hits.
ANNEX_XLSX_URL = (
    "https://taxation-customs.ec.europa.eu/document/download/"
    "1c05d211-80cb-4aaa-8ef0-e08005a95d7e_en"
    "?filename=DVs%20as%20adopted_v20260204%20.xlsx"
)

SEED_DIR = project_root() / "dbt" / "seeds"

# The annex's 121 countries, resolved to ISO3 once, at transcription time. The
# alternative is a fuzzy join at query time, which is the same guesswork with no
# diff to review — and the country list is part of the legal instrument, so
# pinning all of it means an amendment that adds a country fails here with its
# name rather than writing a null ISO3 into the seed. That is exactly what
# happened at the 2026/1740 correction: twelve sheets came back unmapped and the
# script stopped, naming all twelve, instead of writing twelve blank ISO3 codes
# and dropping those countries at the mart's join.
#
# In sheet order, not alphabetical, so this reads against the workbook it mirrors.
#
# Nineteen of these needed a human. Excel truncates a sheet name at 31 characters
# and forbids some punctuation, so a few are the regulation's own name mangled
# ("Congo, Democratic Republic of", "North Korea (Democratic People’" — that one
# is cut mid-word at the 31st character); the rest are the ordinary gap between
# two publishers' country names — the World Bank calls it "Egypt, Arab Rep." and
# the Commission calls it "Egypt". The other 102 match `stg_country` exactly.
#
# Ten of the names changed in the correction with no change to the country: the
# Commission moved to ISO-style long forms ("Russia" -> "Russian Federation",
# "Vietnam" -> "Viet Nam", "Côte d'Ivoire" -> "Ivory Coast"). Liberia and New
# Caledonia are the only two genuinely new entries.
SHEET_TO_ISO3 = {
    "Albania": "ALB",
    "Algeria": "DZA",
    "Angola": "AGO",
    "Argentina": "ARG",
    "Armenia": "ARM",
    "Australia": "AUS",
    "Azerbaijan": "AZE",
    "Bangladesh": "BGD",
    "Bahrain": "BHR",
    "Belarus": "BLR",
    "Benin": "BEN",
    "Bolivia": "BOL",
    "Bosnia and Herzegovina": "BIH",
    "Brazil": "BRA",
    "Brunei": "BRN",
    "Cambodia": "KHM",
    "Cameroon": "CMR",
    "Canada": "CAN",
    "Chile": "CHL",
    "China": "CHN",
    "Colombia": "COL",
    "Congo": "COG",
    "Congo, Democratic Republic of": "COD",
    "Costa Rica": "CRI",
    "Cuba": "CUB",
    "Curaçao": "CUW",
    "Dominican Republic": "DOM",
    "Ecuador": "ECU",
    "Egypt": "EGY",
    "El Salvador": "SLV",
    "Equatorial Guinea": "GNQ",
    "Eritrea": "ERI",
    "Eswatini": "SWZ",
    "Ethiopia": "ETH",
    "Gabon": "GAB",
    "Georgia": "GEO",
    "Ghana": "GHA",
    "Guatemala": "GTM",
    "Haiti": "HTI",
    "Honduras": "HND",
    "Hong Kong": "HKG",
    "India": "IND",
    "Indonesia": "IDN",
    "Iran, Islamic Republic of": "IRN",
    "Iraq": "IRQ",
    "Israel": "ISR",
    "Ivory Coast": "CIV",
    "Jamaica": "JAM",
    "Japan": "JPN",
    "Jordan": "JOR",
    "Kazakhstan": "KAZ",
    "Kenya": "KEN",
    "Korea, Republic of (South Korea": "KOR",
    "Kuwait": "KWT",
    "Kyrgyzstan": "KGZ",
    "Laos": "LAO",
    "Lebanon": "LBN",
    "Liberia": "LBR",
    "Libya": "LBY",
    "Madagascar": "MDG",
    "Malaysia": "MYS",
    "Mali": "MLI",
    "Mauritania": "MRT",
    "Mauritius": "MUS",
    "Mexico": "MEX",
    "Moldova, Republic of": "MDA",
    "Mongolia": "MNG",
    "Montenegro": "MNE",
    "Morocco": "MAR",
    "Mozambique": "MOZ",
    "Myanmar": "MMR",
    "Namibia": "NAM",
    "Nepal": "NPL",
    "New Caledonia and dependencies": "NCL",
    "New Zealand": "NZL",
    "Nicaragua": "NIC",
    "Niger": "NER",
    "Nigeria": "NGA",
    "North Korea (Democratic People’": "PRK",
    "North Macedonia": "MKD",
    "Oman": "OMN",
    "Pakistan": "PAK",
    "Panama": "PAN",
    "Papua New Guinea": "PNG",
    "Paraguay": "PRY",
    "Peru": "PER",
    "Philippines": "PHL",
    "Qatar": "QAT",
    "Russian Federation": "RUS",
    "Rwanda": "RWA",
    "Saudi Arabia": "SAU",
    "Senegal": "SEN",
    "Serbia": "SRB",
    "Sierra Leone": "SLE",
    "Singapore": "SGP",
    "South Africa": "ZAF",
    "Sri Lanka": "LKA",
    "Sudan": "SDN",
    "Suriname": "SUR",
    "Syria": "SYR",
    "Taiwan": "TWN",
    "Tajikistan": "TJK",
    "Tanzania, United Republic of": "TZA",
    "Thailand": "THA",
    "Togo": "TGO",
    "Trinidad and Tobago": "TTO",
    "Tunisia": "TUN",
    "Türkiye": "TUR",
    "Turkmenistan": "TKM",
    "Uganda": "UGA",
    "Ukraine": "UKR",
    "United Arab Emirates": "ARE",
    "United Kingdom": "GBR",
    "United States": "USA",
    "Uruguay": "URY",
    "Uzbekistan": "UZB",
    "Venezuela": "VEN",
    "Viet Nam": "VNM",
    "Yemen": "YEM",
    "Zambia": "ZMB",
    "Zimbabwe": "ZWE",
}

# Column positions on a country sheet, named because they have moved once and a
# bare `row[8]` is what that cost. The February version had **nine** columns —
# direct, indirect, total, then the value *including the mark-up* for 2026, 2027
# and 2028, then the route. The 2026/1740 correction publishes six: the three
# marked-up columns are gone and the route slid from 8 to 5. Nothing failed
# gracefully — `row[8]` raised `IndexError` on the first sheet, which is the good
# outcome; the bad one would have been a layout that still had nine columns
# meaning something else. Hence the width check in `parse_annex`: this script
# refuses a sheet it does not recognise rather than reading the wrong cell.
#
# The mark-up itself is unchanged in law (10/20/30%, fertilisers 1% flat) — the
# Commission simply stopped pre-computing it per row. It is asserted from the
# regulation in the `cbam_markup_schedule` seed now, and applied in
# `fct_cbam_exposure`, because there is no longer a published column to derive it
# from.
COLUMNS = {"cn_code": 0, "description": 1, "direct": 2, "indirect": 3, "total": 4, "route": 5}

# What each of those columns must say it is, checked once per sheet by
# `_check_layout`. Lowercased substrings of the Commission's own headings, so the
# units and parentheses around them can be reworded without breaking the check,
# but a column that *moves* cannot pass.
HEADER_KEYWORDS = {
    "cn_code": "cn code",
    "description": "description",
    "direct": "direct emissions",
    "indirect": "indirect emissions",
    "total": "total emissions",
    "route": "production route",
}

# The annex's catch-all table. It is not a country and gets no ISO3: it is the
# value an importer uses when the sourcing country is unlisted, or is listed with
# a `–` for that good. `fct_cbam_exposure` resolves both cases against it.
FALLBACK_SHEET = "_Other Countries and Territorie"
FALLBACK_LABEL = "Other countries and territories"

# `–` (en dash) is the annex's "no value". `_` and `N/A` also appear; all three
# mean the same thing and all three route to the fallback table.
#
# `see below` is the fourth and it is not a blank at all — it is prose, and it
# appears on exactly two goods: the 4-digit CN headings 3102 and 3105, whose
# numbers live in the subheading rows underneath them. 870 rows (2,610 cells)
# across the workbook. It reached the seed as a null anyway, because the old
# parser returned None for anything it could not read, so the *outcome* was
# right and the reason was luck: the same catch-all would have turned a footnote
# marker or a changed unit into a null just as quietly, and a null here is not
# an absence — `fct_cbam_exposure` reads it as "use the fallback row" and prices
# it. Listing the token is what makes the null a decision. Matched
# case-insensitively, unlike the symbols, because it is a phrase someone typed.
NO_VALUE = {"", "-", "–", "—", "_", "N/A", "n/a", "None"}
NO_VALUE_PHRASES = {"see below"}

# `Annex IV` arrived with the 2026/1740 correction and is deliberately not
# transcribed. It is the single *highest* default value per good with no country
# dimension at all — a different table answering a different question, and the
# circumstances in which a declarant must reach for it instead of the country
# value are set by the articles, not by the annex. Those articles could not be
# confirmed from a primary source here, and inventing a legal trigger is exactly
# what the rest of this script refuses to do. Same posture as Annexes II and III,
# for a different reason: those are excluded on licence, this one on not knowing.
SKIP_SHEETS = {"Overview", "Version History", "Annex IV"}


def _text(cell: object) -> str:
    return "" if cell is None else str(cell).strip()


def _number(cell: object) -> float | None:
    """Parse an annex cell, which may be a float, an int, or a comma decimal.

    Rounded to the three decimals the OJ prints, half away from zero. The mark-up
    columns in the Commission's XLSX are live formulas, so they arrive as binary
    floats — 0.35200000000000004 for a value the regulation states as 0,352 — and
    rounding is what recovers the published figure rather than losing it.

    `round()` will not do: it rounds halves to even, and the mark-up is a
    multiplication that lands on an exact half often enough to matter. China's
    7306 40 80 is 3,300 x 1.1 x ... = 7,7165, which the OJ prints as 7,717 and
    `round()` returns as 7.716. Eleven rows across the seven countries spot-checked
    against the OJ text differed in the third decimal before this changed — small,
    but this column is multiplied by a carbon price and shown as money.

    **An unrecognised token raises rather than returning None**, which is the one
    thing this function must not get wrong. `None` is not an error downstream: it
    is the annex's own "no value here" and `fct_cbam_exposure` reads it as "use
    the fallback row". So a cell this parser merely failed to understand — a
    footnote marker, a stray unit, a thousands-separated figure that leaves two
    separators behind — would not surface as a gap. It would surface as a
    plausible euro figure, correctly computed, attributed to a country the
    regulation never assigned it to. `NO_VALUE` is the list of blanks this
    transcription accepts, and anything outside it is a change in the source that
    a person has to look at. The contract of this script is faithful
    transcription; guessing is the failure, not stopping.
    """
    raw = _text(cell)
    if raw in NO_VALUE or raw.casefold() in NO_VALUE_PHRASES:
        return None
    # Comma is the OJ's *decimal* separator (7,717 is seven point seven one
    # seven), so this is a translation and not a strip. A thousands-separated
    # figure would leave two separators behind and land in the raise below,
    # which is the right answer: this annex does not print one, and reading
    # "1,234,567" as either 1.234567 or 1234567 would be a guess.
    try:
        value = Decimal(repr(float(raw.replace(",", "."))))
    except (ValueError, InvalidOperation) as exc:
        raise ValueError(
            f"unparsable annex cell {raw!r} — if this is a new way of writing "
            f"'no value', add it to NO_VALUE; otherwise the source has changed shape"
        ) from exc
    return float(value.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP))


def _cn_code(cell: object) -> str:
    """Normalise a CN or TARIC code to the spaced form the annex prints.

    Many cells arrive as integers because Excel typed them that way (28142000
    for "2814 20 00"), so the same good reads two different ways depending on
    which country's sheet it is on.

    **10 digits is a TARIC code, and it is new in the 2026/1740 correction.**
    The column is headed "Product CN Code / TARIC Code" and the two are not the
    same length: a CN code is 8 (4+2+2), TARIC adds a further 2. That extra pair
    is what finally distinguishes goods this annex used to publish under one
    number — white clinker and grey clinker were both `2523 10 00` in the
    February version and differed by more than 2x, which is the whole reason
    `good_key` is a slug of (code, description) rather than of the code. They are
    `2523 10 00 10` and `2523 10 00 90` now. The composite key stays: it is what
    kept those two rows apart for the six months the codes could not, and the
    annex still prints 4- and 6-digit headings that no more identify a good than
    the old 8-digit ones did.
    """
    raw = _text(cell)
    if raw.isdigit():
        digits = raw
    else:
        digits = re.sub(r"\s+", "", raw)
        if not digits.isdigit():
            return raw
    # The annex prints 4- and 6-digit headings for whole subheadings alongside
    # the 8-digit CN codes and the 10-digit TARIC ones; all four stay as they are.
    if len(digits) == 10:
        return f"{digits[:4]} {digits[4:6]} {digits[6:8]} {digits[8:]}"
    if len(digits) == 8:
        return f"{digits[:4]} {digits[4:6]} {digits[6:]}"
    if len(digits) == 6:
        return f"{digits[:4]} {digits[4:]}"
    return digits


def _route(cell: object) -> str:
    """The production route indicator, e.g. "(C)" -> "C" and "(C)/(F)" -> "C/F".

    Annex I's footnote maps these to CBAM benchmarks: (C) carbon steel via
    BF/BOF, (E) via scrap/EAF, (K) primary aluminium, (L) secondary, and so on.
    They matter more than they look — the route, not the country's grid, is what
    separates a 0,13 tCO2e/t semi-finished steel from an 8,21.

    Stripping the outer parentheses is not enough: it leaves "C)/(F" for the
    combined routes, which then reads as two different codes in a group-by.
    """
    return "/".join(re.findall(r"\(([A-Z])\)", _text(cell)))


def _slug(text: str, limit: int = 40) -> str:
    ascii_text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    slug = re.sub(r"[^a-z0-9]+", "-", ascii_text.lower()).strip("-")
    return slug[:limit].rstrip("-")


def good_key(cn_code: str, description: str) -> str:
    """A stable, readable key for a (CN code, description) pair.

    A CN code is *not* unique within a country — 2523 10 00 is both white clinker
    and grey clinker, and their default values differ by more than a factor of
    two — so the goods dimension is keyed on the pair. A slug rather than a
    surrogate integer because the seed is regenerated wholesale from each
    amendment, and a renumbered surrogate would rewrite every row of the diff
    when a single CN code is inserted.
    """
    digits = re.sub(r"\s+", "", cn_code)
    return f"{digits}-{_slug(description)}"


def _check_layout(sheet, sheet_name: str) -> None:
    """Fail unless the sheet's header row says what `COLUMNS` assumes it says.

    The width check in the row loop is necessary and not sufficient. It catches a
    layout that *lost* columns — which is what the 2026/1740 correction did, and
    why the previous version blew up loudly on `row[8]` — but the failure that
    actually loses data is a layout that gains one. Every field then still parses,
    one position out, and the only symptom is a column of empty production routes
    or a tonnage that is really a mark-up. Nothing downstream can tell.

    So position is checked against meaning, once per sheet, by reading the
    headings the Commission prints in row 2. Substrings rather than equality: the
    exact wording carries units and parentheses ("Default Value (direct
    emissions) (tCO2eq/tonne of good)") that are cosmetic and have already been
    reworded once, while the words that identify the column have not.
    """
    header = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True), ())
    if len(header) != len(COLUMNS):
        raise ValueError(
            f"{sheet_name!r} header has {len(header)} columns, expected "
            f"{len(COLUMNS)} — the annex layout has changed; see COLUMNS"
        )
    for name, index in COLUMNS.items():
        expected = HEADER_KEYWORDS[name]
        found = _text(header[index]).casefold()
        if expected not in found:
            raise ValueError(
                f"{sheet_name!r} column {index} reads {_text(header[index])!r}, "
                f"expected it to mention {expected!r} — the annex columns have "
                f"moved and COLUMNS is now pointing at the wrong cell"
            )


def parse_annex(xlsx_path: Path) -> tuple[list[dict], list[dict]]:
    """Return (goods dimension rows, default-value rows) from the annex workbook."""
    try:
        import openpyxl
    except ModuleNotFoundError:  # pragma: no cover - dev-only dependency
        sys.exit("openpyxl is required: uv run --with openpyxl python -m scripts.build_cbam_seeds")

    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    goods: dict[str, dict] = {}
    values: list[dict] = []

    for sheet_name in workbook.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        is_fallback = sheet_name == FALLBACK_SHEET
        if is_fallback:
            country, iso3 = FALLBACK_LABEL, ""
        else:
            country = sheet_name
            iso3 = SHEET_TO_ISO3.get(sheet_name, "")

        _check_layout(workbook[sheet_name], sheet_name)

        product_group = ""
        for row in workbook[sheet_name].iter_rows(min_row=3, values_only=True):
            code = _text(row[0]) if row else ""
            if not code:
                continue
            # Exactly, not "at least". A row with *more* columns than expected is
            # the dangerous direction — every field would still parse, one
            # position out — and `<` waved it through.
            if len(row) != len(COLUMNS):
                raise ValueError(
                    f"{sheet_name!r} row {code!r} has {len(row)} columns, expected "
                    f"{len(COLUMNS)} — the annex layout has changed"
                )
            description = _text(row[1])
            if not description:
                # A group banner ("Cement", "Iron and steel").
                product_group = code
                continue

            cn_code = _cn_code(row[0])
            key = good_key(cn_code, description)
            goods.setdefault(
                key,
                {
                    "good_key": key,
                    "product_group": product_group,
                    "cn_code": cn_code,
                    "goods_description": description,
                },
            )

            values.append(
                {
                    "country_or_territory": country,
                    "country_iso3": iso3,
                    "good_key": key,
                    "default_direct_t_co2e_per_t": _number(row[COLUMNS["direct"]]),
                    "default_indirect_t_co2e_per_t": _number(row[COLUMNS["indirect"]]),
                    "default_total_t_co2e_per_t": _number(row[COLUMNS["total"]]),
                    "production_route_code": _route(row[COLUMNS["route"]]),
                }
            )

    return sorted(
        goods.values(), key=lambda g: (g["product_group"], g["cn_code"], g["good_key"])
    ), values


def _write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({k: ("" if row[k] is None else row[k]) for k in fieldnames})
    print(f"wrote {path.relative_to(project_root())}: {len(rows):,} rows")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, help="local copy of the annex workbook")
    args = parser.parse_args()

    xlsx_path = args.xlsx
    if xlsx_path is None:
        xlsx_path = project_root() / "data" / "cbam_annex_i.xlsx"
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        print(f"downloading {ANNEX_XLSX_URL}")
        with urlopen(ANNEX_XLSX_URL) as response:
            xlsx_path.write_bytes(response.read())

    goods, values = parse_annex(xlsx_path)

    unmapped = sorted(
        {v["country_or_territory"] for v in values if not v["country_iso3"]} - {FALLBACK_LABEL}
    )
    if unmapped:
        sys.exit(f"no ISO3 for {len(unmapped)} territories, add them to SHEET_TO_ISO3: {unmapped}")

    _write_csv(
        SEED_DIR / "cbam_goods.csv",
        goods,
        ["good_key", "product_group", "cn_code", "goods_description"],
    )
    _write_csv(
        SEED_DIR / "cbam_default_values.csv",
        values,
        [
            "country_or_territory",
            "country_iso3",
            "good_key",
            "default_direct_t_co2e_per_t",
            "default_indirect_t_co2e_per_t",
            "default_total_t_co2e_per_t",
            "production_route_code",
        ],
    )


if __name__ == "__main__":
    main()
